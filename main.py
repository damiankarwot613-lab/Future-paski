# ==========================================
# PASKI FUTURE - ANDROID 14–16 STABLE
# SAF + ANDROIDX
# ==========================================

import json
import threading
from pathlib import Path

from kivy.app import App
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.utils import platform
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.progressbar import ProgressBar
from kivy.uix.screenmanager import ScreenManager, Screen


APP_TITLE = "Paski Future Stable"
PREVIEW_ROWS = 5


class Home(Screen):
    pass


class Table(Screen):
    pass


class PaskiApp(App):

    def build(self):

        self.full_data = []
        self.filtered_data = []
        self.current_file = None
        self.export_tree_uri = None

        self.config_path = Path(self.user_data_dir) / "config.json"

        self.sm = ScreenManager()
        self.home = Home(name="home")
        self.table = Table(name="table")

    # OPÓŹNIONA INICJALIZACJA
    Clock.schedule_once(self.finish_init, 1)

    return self.sm

    def finish_init(self, dt):

        self.load_config()

        self.build_home()
        self.build_table()

        self.sm.add_widget(self.home)
        self.sm.add_widget(self.table)

        self.sm.current = "home"

    # ================= CONFIG =================

    def load_config(self):
        if self.config_path.exists():
            with open(self.config_path, "r") as f:
                data = json.load(f)
                self.export_tree_uri = data.get("export_tree_uri")

    def save_config(self):
        with open(self.config_path, "w") as f:
            json.dump({"export_tree_uri": self.export_tree_uri}, f)

    # ================= HOME =================

    def build_home(self):
        layout = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10))

        layout.add_widget(Label(text=APP_TITLE, size_hint=(1, 0.15)))

        pick_file = Button(text="Wybierz plik Excel")
        pick_folder = Button(text="Wybierz folder eksportu")
        load_btn = Button(text="Wczytaj pełne dane")

        self.preview_label = Label(
            text="Brak pliku",
            halign="left",
            valign="top"
        )
        self.preview_label.bind(size=self.update_text_size)

        pick_file.bind(on_press=self.open_excel_picker)
        pick_folder.bind(on_press=self.open_folder_picker)
        load_btn.bind(on_press=self.load_full_excel)

        layout.add_widget(pick_file)
        layout.add_widget(pick_folder)
        layout.add_widget(load_btn)
        layout.add_widget(self.preview_label)

        self.home.add_widget(layout)

    def update_text_size(self, *_):
        self.preview_label.text_size = self.preview_label.size

    # ================= PICKERS =================

    def open_excel_picker(self, _):
        from jnius import autoclass
        from android import activity

        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        Intent = autoclass("android.content.Intent")

        intent = Intent(Intent.ACTION_OPEN_DOCUMENT)
        intent.setType("*/*")
        intent.addCategory(Intent.CATEGORY_OPENABLE)
        intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)

        PythonActivity.mActivity.startActivityForResult(intent, 1001)
        activity.bind(on_activity_result=self.on_activity_result)

    def open_folder_picker(self, _):
        from jnius import autoclass
        from android import activity

        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        Intent = autoclass("android.content.Intent")

        intent = Intent(Intent.ACTION_OPEN_DOCUMENT_TREE)
        intent.addFlags(
            Intent.FLAG_GRANT_READ_URI_PERMISSION |
            Intent.FLAG_GRANT_WRITE_URI_PERMISSION |
            Intent.FLAG_GRANT_PERSISTABLE_URI_PERMISSION
        )

        PythonActivity.mActivity.startActivityForResult(intent, 2001)
        activity.bind(on_activity_result=self.on_activity_result)

    def on_activity_result(self, request_code, result_code, intent):
        from android import activity
        activity.unbind(on_activity_result=self.on_activity_result)

        if result_code != -1 or not intent:
            return

        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        resolver = PythonActivity.mActivity.getContentResolver()
        uri = intent.getData()

        if request_code == 1001:
            input_stream = resolver.openInputStream(uri)
            local_file = Path(self.user_data_dir) / "selected.xlsx"

            with open(local_file, "wb") as f:
                f.write(input_stream.read())

            input_stream.close()
            self.current_file = local_file
            self.show_preview()

        elif request_code == 2001:
            flags = (intent.getFlags()
                     & (intent.FLAG_GRANT_READ_URI_PERMISSION |
                        intent.FLAG_GRANT_WRITE_URI_PERMISSION))

            resolver.takePersistableUriPermission(uri, flags)
            self.export_tree_uri = uri.toString()
            self.save_config()
            self.popup("Folder zapisany")

    # ================= PREVIEW =================

    def show_preview(self):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(self.current_file), data_only=True)
            sheet = wb.active

            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= PREVIEW_ROWS:
                    break
                rows.append(" | ".join("" if v is None else str(v) for v in row))

            wb.close()
            self.preview_label.text = "\n".join(rows)

        except Exception as e:
            self.preview_label.text = f"Błąd: {e}"

    # ================= LOAD =================

    def load_full_excel(self, _):
        if not self.current_file:
            self.popup("Najpierw wybierz plik")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.current_file), data_only=True)
            sheet = wb.active

            self.full_data = [
                ["" if v is None else str(v) for v in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            wb.close()

        except Exception as e:
            self.popup(f"Błąd: {e}")
            return

        self.filtered_data = self.full_data
        self.display_table()
        self.sm.current = "table"

    # ================= TABLE =================

    def build_table(self):
        layout = BoxLayout(orientation="vertical")

        top = BoxLayout(size_hint=(1, 0.1))
        self.search = TextInput(hint_text="Szukaj...", multiline=False)
        self.search.bind(text=self.filter_data)

        export_btn = Button(text="Eksport")
        back_btn = Button(text="Powrót")

        export_btn.bind(on_press=self.export_files)
        back_btn.bind(on_press=lambda x: setattr(self.sm, "current", "home"))

        top.add_widget(self.search)
        top.add_widget(export_btn)
        top.add_widget(back_btn)

        self.scroll = ScrollView()
        self.grid = GridLayout(size_hint_y=None)
        self.grid.bind(minimum_height=self.grid.setter("height"))
        self.scroll.add_widget(self.grid)

        self.progress = ProgressBar(max=100, size_hint=(1, 0.05))

        layout.add_widget(top)
        layout.add_widget(self.scroll)
        layout.add_widget(self.progress)

        self.table.add_widget(layout)

    def display_table(self):
        self.grid.clear_widgets()

        if not self.filtered_data:
            return

        self.grid.cols = len(self.filtered_data[0])

        for row in self.filtered_data:
            for cell in row:
                self.grid.add_widget(
                    Label(text=str(cell), size_hint_y=None, height=40)
                )

    def filter_data(self, _, value):
        value = value.lower()
        self.filtered_data = [
            row for row in self.full_data
            if any(value in str(cell).lower() for cell in row)
        ]
        self.display_table()

    # ================= EXPORT =================

    def export_files(self, _):
        if not self.export_tree_uri:
            self.popup("Wybierz folder eksportu")
            return

        if len(self.filtered_data) <= 1:
            self.popup("Brak danych")
            return

        threading.Thread(target=self.export_thread).start()

    def export_thread(self):
        from openpyxl import Workbook
        from jnius import autoclass

        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        Uri = autoclass("android.net.Uri")
        DocumentFile = autoclass("androidx.documentfile.provider.DocumentFile")

        resolver = PythonActivity.mActivity.getContentResolver()
        tree_uri = Uri.parse(self.export_tree_uri)
        tree = DocumentFile.fromTreeUri(PythonActivity.mActivity, tree_uri)

        header = self.full_data[0]
        rows = self.filtered_data[1:]
        total = len(rows)

        for i, row in enumerate(rows):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            ws.append(row)

            file = tree.createFile(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"plik_{i+1}.xlsx"
            )

            output = resolver.openOutputStream(file.getUri())
            wb.save(output)
            wb.close()
            output.close()

            percent = int(((i+1) / total) * 100)
            Clock.schedule_once(lambda dt, p=percent:
                                setattr(self.progress, "value", p))

        Clock.schedule_once(lambda dt:
                            self.popup("Eksport zakończony"))

    # ================= POPUP =================

    def popup(self, message):
        Popup(
            title="Info",
            content=Label(text=str(message)),
            size_hint=(0.8, 0.4)
        ).open()


if __name__ == "__main__":
    PaskiApp().run()
